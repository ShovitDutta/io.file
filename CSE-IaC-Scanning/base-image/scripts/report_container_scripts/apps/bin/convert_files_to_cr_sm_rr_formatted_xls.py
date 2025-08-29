import openpyxl
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from csv import reader
import sys

def adjust_width(ws,snm):
    """
    Adjust width of the columns
    @param ws: worksheet
    @return: None
    """

    def is_merged_horizontally(cell):
        """
        Checks if cell is merged horizontally with an another cell
        @param cell: cell to check
        @return: True if cell is merged horizontally with an another cell, else False
        """
        cell_coor = cell.coordinate
        if cell_coor not in ws.merged_cells:
            return False
        for rng in ws.merged_cells.ranges:
            if cell_coor in rng and len(list(rng.cols)) > 1:
                return True
        return False

    for col_number, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_number)

        max_length = max(
            len(str(cell.value or "")) for cell in col if not is_merged_horizontally(cell)
        )
        adjusted_width = (max_length + 2) * 0.95
        ws.column_dimensions[col_letter].width = adjusted_width
        #ws.column_dimensions[col_letter].number_format = "########"
        
def process_cell_format(col,str_value,format):
    """
    format number columns
    @param ws: col
    @return: None
    """
    for cell in col:
        if len(str(cell.value or "")) > 0:
           if cell.value != str_value:
              cell.value=int(cell.value)
              cell.number_format=format


# Define variable to load the dataframe
wb = openpyxl.Workbook()
sheet_array=[ "sheet0", "sheet1","sheet2","sheet3","sheet4","sheet5","sheet6","sheet7","sheet8","sheet9","sheet10","sheet11","sheet12" ]
sheet_array_nm=[ "sheet0","sheet1", "Coverage_Report","Scan_metrics(kics)","Scan_metrics(snyk)","Weekly metrics by rules","Rules and Resources","list of kics rules" ]

n = len(sys.argv)
# open demo.csv file in read mode
for i in range(2, n):
      with open(sys.argv[i], 'r') as readObj:

    # pass the file object to reader() to get the reader object
           csvReader = reader(readObj)

           my_sheet=sheet_array[i]
           my_sheet_nm=sheet_array_nm[i]
           my_sheet=wb.create_sheet(my_sheet_nm)
           my_sheet.number_format = 'General'

    # Iterate over each row in the csv using reader object
           for row in csvReader:
        # row variable is a list that represents a row in csv
               my_sheet.append(row)
               adjust_width(my_sheet,my_sheet_nm)
           #if my_sheet_nm == "Coverage_Report":
              #for col_number, col in enumerate(my_sheet.columns, start=1):
                   #col_letter = get_column_letter(col_number)
                   #if col_letter == "B":
                      #process_cell_format(col,"Pipelines","000")
                   #if col_letter == "C":
                      #process_cell_format(col,"StandardRules","00000")
                   #if col_letter == "D":
                      #process_cell_format(col,"Custom Rules","00000")
           #if my_sheet_nm == "Scan_metrics(snyk)":
           #   for col_number, col in enumerate(my_sheet.columns, start=1):
           #        col_letter = get_column_letter(col_number)
           #        if col_letter == "B":
           #           process_cell_format(col,"WeekofYear","00")
           #if my_sheet_nm == "Scan_metrics(kics)":
           #   for col_number, col in enumerate(my_sheet.columns, start=1):
           #        col_letter = get_column_letter(col_number)
           #        if col_letter == "B":
           #           process_cell_format(col,"WeekofYear","0000")
           #if my_sheet_nm == "Weekly metrics by rules(snyk)":
              #for col_number, col in enumerate(my_sheet.columns, start=1):
                   #col_letter = get_column_letter(col_number)
                   #if col_letter == "G":
                      #process_cell_format(col,"Week of Year","000")
                   #if col_letter == "H":
                      #process_cell_format(col,"Rule Count","00000")
           #if my_sheet_nm == "Weekly metrics by rules(kics)":
              #for col_number, col in enumerate(my_sheet.columns, start=1):
                   #col_letter = get_column_letter(col_number)
                   #if col_letter == "G":
                      #process_cell_format(col,"Week of Year","000")
                   #if col_letter == "H":
                      #process_cell_format(col,"Rule Count","00000")

                 
wb.remove(wb['Sheet'])
wb.save(sys.argv[1])
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from csv import reader
import sys

def adjust_width(ws,snm):
    """
    Adjust width of the columns
    @param ws: worksheet
    @return: None
    """

    def is_merged_horizontally(cell):
        """
        Checks if cell is merged horizontally with an another cell
        @param cell: cell to check
        @return: True if cell is merged horizontally with an another cell, else False
        """
        cell_coor = cell.coordinate
        if cell_coor not in ws.merged_cells:
            return False
        for rng in ws.merged_cells.ranges:
            if cell_coor in rng and len(list(rng.cols)) > 1:
                return True
        return False

    for col_number, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_number)

        max_length = max(
            len(str(cell.value or "")) for cell in col if not is_merged_horizontally(cell)
        )
        adjusted_width = (max_length + 2) * 0.95
        ws.column_dimensions[col_letter].width = adjusted_width
        #ws.column_dimensions[col_letter].number_format = "########"
        
def process_cell_format(col,str_value,format):
    """
    format number columns
    @param ws: col
    @return: None
    """
    for cell in col:
        if len(str(cell.value or "")) > 0:
           if cell.value != str_value:
              cell.value=int(cell.value)
              cell.number_format=format


# Define variable to load the dataframe
wb = openpyxl.Workbook()
sheet_array=[ "sheet0", "sheet1","sheet2","sheet3","sheet4","sheet5","sheet6","sheet7","sheet8","sheet9","sheet10","sheet11","sheet12" ]
sheet_array_nm=[ "sheet0","sheet1", "Coverage_Report","Scan_metrics(kics)","Scan_metrics(snyk)","Weekly metrics by rules","Rules and Resources","list of kics rules" ]

n = len(sys.argv)
# open demo.csv file in read mode
for i in range(2, n):
      with open(sys.argv[i], 'r') as readObj:

    # pass the file object to reader() to get the reader object
           csvReader = reader(readObj)

           my_sheet=sheet_array[i]
           my_sheet_nm=sheet_array_nm[i]
           my_sheet=wb.create_sheet(my_sheet_nm)
           my_sheet.number_format = 'General'

    # Iterate over each row in the csv using reader object
           for row in csvReader:
        # row variable is a list that represents a row in csv
               my_sheet.append(row)
               adjust_width(my_sheet,my_sheet_nm)
           #if my_sheet_nm == "Coverage_Report":
              #for col_number, col in enumerate(my_sheet.columns, start=1):
                   #col_letter = get_column_letter(col_number)
                   #if col_letter == "B":
                      #process_cell_format(col,"Pipelines","000")
                   #if col_letter == "C":
                      #process_cell_format(col,"StandardRules","00000")
                   #if col_letter == "D":
                      #process_cell_format(col,"Custom Rules","00000")
           #if my_sheet_nm == "Scan_metrics(snyk)":
           #   for col_number, col in enumerate(my_sheet.columns, start=1):
           #        col_letter = get_column_letter(col_number)
           #        if col_letter == "B":
           #           process_cell_format(col,"WeekofYear","00")
           #if my_sheet_nm == "Scan_metrics(kics)":
           #   for col_number, col in enumerate(my_sheet.columns, start=1):
           #        col_letter = get_column_letter(col_number)
           #        if col_letter == "B":
           #           process_cell_format(col,"WeekofYear","0000")
           #if my_sheet_nm == "Weekly metrics by rules(snyk)":
              #for col_number, col in enumerate(my_sheet.columns, start=1):
                   #col_letter = get_column_letter(col_number)
                   #if col_letter == "G":
                      #process_cell_format(col,"Week of Year","000")
                   #if col_letter == "H":
                      #process_cell_format(col,"Rule Count","00000")
           #if my_sheet_nm == "Weekly metrics by rules(kics)":
              #for col_number, col in enumerate(my_sheet.columns, start=1):
                   #col_letter = get_column_letter(col_number)
                   #if col_letter == "G":
                      #process_cell_format(col,"Week of Year","000")
                   #if col_letter == "H":
                      #process_cell_format(col,"Rule Count","00000")

                 
wb.remove(wb['Sheet'])
wb.save(sys.argv[1])
