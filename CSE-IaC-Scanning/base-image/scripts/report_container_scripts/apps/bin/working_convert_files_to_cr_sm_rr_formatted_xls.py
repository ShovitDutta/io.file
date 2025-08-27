import openpyxl
from openpyxl.utils import get_column_letter
from csv import reader
import sys

def adjust_width(ws):
    """
    Adjust width of the columns
    @param ws: worksheet
    @return: None
    """

    def is_merged_horizontally(cell):
        """
        Checks if cell is merged horizontally with an another cell
        @param cell: cell to check
        @return: True if cell is merged horizontally with an another cell, else False
        """
        cell_coor = cell.coordinate
        if cell_coor not in ws.merged_cells:
            return False
        for rng in ws.merged_cells.ranges:
            if cell_coor in rng and len(list(rng.cols)) > 1:
                return True
        return False

    for col_number, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_number)

        max_length = max(
            len(str(cell.value or "")) for cell in col if not is_merged_horizontally(cell)
        )
        adjusted_width = (max_length + 2) * 0.95
        ws.column_dimensions[col_letter].width = adjusted_width
        


# Define variable to load the dataframe
wb = openpyxl.Workbook()
sheet_array=[ "sheet0", "sheet1","sheet2","sheet3","sheet4","sheet5","sheet6","sheet7" ]
sheet_array_nm=[ "sheet0","sheet1", "Coverage_Report","Scan_metrics(snyk)","Scan_metrics(kics)","Weekly metrics by rules(snyk)","Weekly metrics by rules(kics)" ]

n = len(sys.argv)
# open demo.csv file in read mode
for i in range(2, n):
      with open(sys.argv[i], 'r') as readObj:

    # pass the file object to reader() to get the reader object
           csvReader = reader(readObj)

           my_sheet=sheet_array[i]
           my_sheet_nm=sheet_array_nm[i]
           my_sheet=wb.create_sheet(my_sheet_nm)
    # Iterate over each row in the csv using reader object
           for row in csvReader:
        # row variable is a list that represents a row in csv
              my_sheet.append(row)
              adjust_width(my_sheet)
wb.remove(wb['Sheet'])
wb.save(sys.argv[1])
